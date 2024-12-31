from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

AllPageContent = []
ShopName = []
ShopItems = []
ShopStar = []
ShopOutDay = []
ShopResponse = []
ProductsManu = []
ProductsName = []
ProductsDiscount = []
ProductsPrice = []
ProductsStar = []
ProductsHot = []
ProductsTWMade = []
ProductsSafety = []

Url = "https://www.pcone.com.tw/product/603#ref=d_nav"
Driver = webdriver.Edge()
Driver.get(Url)
Driver.implicitly_wait(6)

def ClassLoad():
    for PageContent in AllPageContent:
        Html = BeautifulSoup(PageContent, 'html.parser')
        AllProducts = Html.find(class_ = "dynamic-products product-list")
        Chunk = AllProducts.find_all(class_ = "product-list-item")
        RunIndex = 0
        for SubChunk in Chunk:
            RunIndex += 1
            if (RunIndex > 110):
                RunIndex = 0
                time.sleep(120)
            Driver.get(f"https://www.pcone.com.tw{SubChunk.get('href')}")
            Driver.find_element(By.TAG_NAME ,'body').send_keys(Keys.END)
            time.sleep(1)
            ProductContent = Driver.page_source # 取得頁面內容
            ProductHtml = BeautifulSoup(ProductContent, 'html.parser')
            try:
                print(f"店家名稱：{ProductHtml.find(class_ = 'merchant-name').get_text()}")
            except AttributeError:
                continue
            ShopName_M = ProductHtml.find(class_ = 'merchant-name').get_text()
            details = ProductHtml.find('div', class_ = 'details')
            ShopItems_M = details.find_all('div', class_ = 'count')[0].text
            ShopStar_M = details.find_all('div', class_ = 'count')[1].text
            ShopOutDay_M = details.find_all('div', class_ = 'count')[2].text
            ShopResponse_M = details.find_all('div', class_ = 'count')[3].text
            ProductsName_M = ProductHtml.find(class_ = 'name').get_text()
            try:
                ProductsDiscount_M = ProductHtml.find(class_ = 'tag circle').get_text()
            except AttributeError:
                ProductsDiscount_M = "無特價"
            ProductsPrice_M = ProductHtml.find(class_ = 'display-price').get_text()
            details = ProductHtml.find('div', class_='review-info')
            if details.find_all('div')[1].text:
                ProductsHot_M = details.find_all('div')[1].text.split( )[0]
            else: 
                ProductsHot_M = ""
            if details.find_all('div')[0].text:
                ProductsStar_M = details.find_all('div')[0].text.split( )[0]
            else:
                ProductsStar_M = ""
            ShopName.append(ShopName_M) # 店家名稱
            ShopItems.append(ShopItems_M) # 店家商品數量
            ShopStar.append(ShopStar_M) # 店家評價
            ShopOutDay.append(ShopOutDay_M) # 店家出貨天數
            ShopResponse.append(ShopResponse_M) # 店家回覆率
            ProductsName.append(ProductsName_M) # 商品名稱
            ProductsDiscount.append(ProductsDiscount_M[:-1]) # 商品折扣
            ProductsPrice.append(ProductsPrice_M[1:]) # 商品價錢
            ProductsHot.append(ProductsHot_M) # 商品銷量多寡
            ProductsStar.append(ProductsStar_M) # 商品評價
            # 是否為台灣製
            if ("台灣製" in ProductHtml.find(class_ = 'name').get_text()):
                ProductsTWMade.append(1)
            else:
                ProductsTWMade.append(0)

            # 是否安全
            if ("安全" in ProductHtml.find('article', class_='markdown-body markdown-body-img editormd-html-preview') or "安全" in ProductHtml.find(class_ = 'name').get_text()):
                ProductsSafety.append(1)
            else:
                ProductsSafety.append(0)

# def SearchLoad():
#     for PageContent in AllPageContent:
#         Html = BeautifulSoup(PageContent, 'html.parser')
#         Chunk = Html.find_all(class_ = "chunk-commodities")
#         for SubChunk in Chunk:
#             for Info in SubChunk.find_all('a'):
#                 Driver.get(Info.get('href'))
#                 ProductContent = Driver.page_source # 取得頁面內容
#                 ProductHtml = BeautifulSoup(ProductContent, 'html.parser')
#                 try:
#                     print(f"店家名稱：{ProductHtml.find(class_ = 'merchant-name').get_text()}")
#                     ShopName_M = ProductHtml.find(class_ = 'merchant-name').get_text()
#                     details = ProductHtml.find('div', class_ = 'details')
#                     ShopItems_M = details.find_all('div', class_ = 'count')[0].text
#                     ShopStar_M = details.find_all('div', class_ = 'count')[1].text
#                     ShopOutDay_M = details.find_all('div', class_ = 'count')[2].text
#                     ShopResponse_M = details.find_all('div', class_ = 'count')[3].text
#                     ProductsName_M = ProductHtml.find(class_ = 'name').get_text()
#                     ProductsDiscount_M = ProductHtml.find(class_ = 'tag circle').get_text()
#                     ProductsPrice_M = ProductHtml.find(class_ = 'display-price').get_text()
#                     details = ProductHtml.find('div', class_='review-info')
#                     if details.find_all('div')[1].text:
#                         ProductsHot_M = details.find_all('div')[1].text.split( )[0]
#                     else: 
#                         ProductsHot_M = ""
#                     if details.find_all('div')[0].text:
#                         ProductsStar_M = details.find_all('div')[0].text.split( )[0]
#                     else:
#                         ProductsStar_M = ""
#                     ShopName.append(ShopName_M) # 店家名稱
#                     ShopItems.append(ShopItems_M) # 店家商品數量
#                     ShopStar.append(ShopStar_M) # 店家評價
#                     ShopOutDay.append(ShopOutDay_M) # 店家出貨天數
#                     ShopResponse.append(ShopResponse_M) # 店家回覆率
#                     ProductsName.append(ProductsName_M) # 商品名稱
#                     ProductsDiscount.append(ProductsDiscount_M[:-1]) # 商品折扣
#                     ProductsPrice.append(ProductsPrice_M[1:]) # 商品價錢
#                     ProductsHot.append(ProductsHot_M) # 商品銷量多寡
#                     ProductsStar.append(ProductsStar_M) # 商品評價
#                     # 是否為台灣製
#                     if ("台灣製" in ProductHtml.find(class_ = 'name').get_text()):
#                         ProductsTWMade.append(1)
#                     else:
#                         ProductsTWMade.append(0)

#                     # 是否安全
#                     if ("安全" in ProductHtml.find('article', class_='markdown-body markdown-body-img editormd-html-preview')):
#                         ProductsSafety.append(1)
#                     else:
#                         ProductsSafety.append(0)
                    
#                 except AttributeError:
#                     print("NULL")

# Search 頁面
# def NewOnePage(page):
#     for index in range(page):
#         AllPageContent.append(Driver.page_source) # 取得頁面內容
#         print(len(AllPageContent))
#         if (index+1 < page):
#             if (index+1 > 3):
#                 Driver.find_element(By.XPATH, '//*[@id="__next"]/div[1]/div/div[3]/div[2]/div[2]/div[2]/a[8]').click()
#             else:
#                 Driver.find_element(By.XPATH, '//*[@id="__next"]/div[1]/div/div[3]/div[2]/div[2]/div[2]/a[9]').click()

def ClassificationPage(scroll):
    for index in range(scroll):
        Driver.find_element(By.TAG_NAME ,'body').send_keys(Keys.END)
        time.sleep(2)
    AllPageContent.append(Driver.page_source) # 取得頁面內容

def CreateWriteExcel():
    # 建立excel檔並寫入
    OutputExcel = "shoplist.xlsx"
    try:
        Wb = openpyxl.load_workbook(OutputExcel)
        Sheet = Wb["商品清單"]
    except FileNotFoundError:
        Wb = openpyxl.Workbook()
        Wb.create_sheet("商品清單", 0)
        Sheet = Wb["商品清單"]
        for index, value in [
            ("A", "店家名稱"),
            ("B", "店家商品數量"),
            ("C", "店家評價"),
            ("D", "店家出貨天數"),
            ("E", "店家回覆率"),
            ("F", "商品名稱"),
            ("G", "商品折扣"),
            ("H", "商品價錢"),
            ("I", "商品銷量多寡"),
            ("J", "商品評價"),
            ("K", "是否為台灣製"),
            ("L", "是否提及安全"),
            ("M", "是否為高價品")
        ]:
            Sheet[f"{index}1"].value = value
        Wb.save(OutputExcel)
    print(len(ProductsDiscount))
    for Row, value1 in enumerate(ShopName):
        for Column, value2 in [
            ("A", ShopName[Row]),
            ("B", ShopItems[Row]),
            ("C", ShopStar[Row]),
            ("D", ShopOutDay[Row]),
            ("E", ShopResponse[Row]),
            ("F", ProductsName[Row]),
            ("G", ProductsDiscount[Row]),
            ("H", ProductsPrice[Row]),
            ("I", ProductsHot[Row]),
            ("J", ProductsStar[Row]),
            ("K", ProductsTWMade[Row]),
            ("L", ProductsSafety[Row])
        ]:
            Sheet[f"{Column}{Row + 2}"].value = value2

    Wb.save(OutputExcel)
    Wb.close()

ClassificationPage(25)
ClassLoad()
CreateWriteExcel()

print(f"店家數量：{len(list(set(ShopName)))}")
print(f"商品筆數：{len(ProductsName)}")
Driver.quit()